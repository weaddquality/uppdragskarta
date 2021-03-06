# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import datetime
#Output file
wbo = Workbook()
wso = wbo.active
#Assignement Woorkbook
try:
	wba = load_workbook(filename = 'Cinode_CurrentAssignments_Export.xlsx')
except FileNotFoundError:
	print("Fil med uppdrag saknas")
	sys.exit()
#Customer addresses
try:
	wbc = load_workbook(filename = 'Cinode_Customer_Addresses_Export.xlsx')
except FileNotFoundError:
	print("Fil med adresser saknas")
	sys.exit()
#WorkBook Sheet
wsa = wba.active
wsc = wbc.active
customers = [] #list for temporarily storing customer names
consultant_column = {}
customer_address_exist = 0
#Headers
wso.cell(row=1,column=1).value = 'Kund'
wso.cell(row=1,column=2).value = 'Gata'
wso.cell(row=1,column=3).value = 'Postnummer'
wso.cell(row=1,column=4).value = 'Stad'
# insert customer names ones
count = 2 # row to start inserting values
for i in range(2,wsa.max_row+1):
        if customer_address_exist == 0 and i != 2:
                print (temp_customer + "- Adress saknas")
        customer_address_exist = 0
        temp_customer = wsa.cell(row=i,column=8).value
        if temp_customer in customers or temp_customer == "ADDQ":
                customer_address_exist = 1
                pass
        else:
                customers.append(temp_customer)
                wso.cell(row=count,column=1).value = temp_customer #insert customer name in output file
                for j in range(2,wsc.max_row+1):
                        consultant_column[temp_customer] = 5 # The first consultant should be in column 5
                        #if wsc.cell(row=j,column=2).value == temp_customer:
                        if wsc.cell(row=j,column=2).value == temp_customer and str(wsc.cell(row=j,column=9).value) == 'Street':
                                customer_address_exist = 1
                                wso.cell(row=count,column=2).value = wsc.cell(row=j,column=4).value #Customer addresse
                                wso.cell(row=count,column=3).value = wsc.cell(row=j,column=6).value #Customer postal code
                                wso.cell(row=count,column=4).value = wsc.cell(row=j,column=7).value #Customer city
                                if wsc.cell(row=j,column=4).value is None or wsc.cell(row=j,column=6).value is None or wsc.cell(row=j,column=7).value is None:
                                        temp_saknas = ""
                                        if wsc.cell(row=j,column=4).value is None:
                                                temp_saknas = temp_saknas + "Gatunamn"
                                        if wsc.cell(row=j,column=6).value is None:
                                                temp_saknas = temp_saknas + " Postnummer"
                                        if wsc.cell(row=j,column=7).value is None:
                                                temp_saknas = temp_saknas + " Stad"
                                        print (temp_customer +"- Följande adressdata saknas:" + temp_saknas)
                                break
                count = count + 1
#Insert consultant name
consultant = []
for cust_row in range(1, wso.max_row+1): # iterate over output file
	consultant.clear()
	for assign_row in range(1,wsa.max_row+1): #iterate over assignment file
		if wsa.cell(row=assign_row,column=8).value == wso.cell(row=cust_row,column=1).value: #compare customer name in assignment with output file
			temp_consultant = wsa.cell(row=assign_row,column=1).value
			temp_status = wsa.cell(row=assign_row,column=4).value
			if temp_consultant in consultant:
				pass
			elif temp_status != "Tillsatt": #Tar endast med tillsatta roller, inte de som blivit avböjda eller avböjts
				pass
			else:
				#insert consultant name in output file in column 'consultant_column'
				wso.cell(row=cust_row, column=consultant_column[wso.cell(row=cust_row,column=1).value]).value = wsa.cell(row=assign_row,column=1).value
				consultant_column[wso.cell(row=cust_row,column=1).value] += 1
				consultant.append(temp_consultant)
tempday = datetime.date.today().strftime("%Y-%m-%d") #
wbo.save('output_' + tempday + '.xlsx')
print("Output file created")